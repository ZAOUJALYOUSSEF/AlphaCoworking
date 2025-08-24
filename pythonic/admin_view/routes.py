from flask import render_template, request, flash, jsonify, redirect, url_for, make_response
from flask_login import login_required, current_user
from pythonic import db
from pythonic.models import Booking, ContactMessage
from sqlalchemy import func, extract, cast, Date, or_, and_
from datetime import datetime, timedelta, date, time
from pythonic.admin_view import admin_stats
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from flask import Blueprint, request, send_file


@admin_stats.route("/dashboard/admin_stats")
@login_required
def stats_dashboard():
    # Récupérer les paramètres de filtre
    filter_type = request.args.get('filter_type', 'day')
    space_type_filter = request.args.get('space_type', 'all')
    today = datetime.now().date()

    # Récupérer la liste des types d'espace disponibles
    available_space_types = db.session.query(Booking.space_type.distinct()).order_by(Booking.space_type).all()
    available_space_types = [space_type[0] for space_type in available_space_types if space_type[0] is not None]

    # Gestion des dates selon le type de filtre
    if filter_type == 'day':
        selected_date = request.args.get('selected_date', today.strftime('%Y-%m-%d'))
        try:
            filter_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            start_date = filter_date
            end_date = filter_date
            title_suffix = f"du {filter_date.strftime('%d/%m/%Y')}"
        except:
            filter_date = today
            start_date = today
            end_date = today
            title_suffix = f"du {today.strftime('%d/%m/%Y')}"
    elif filter_type == 'month':
        selected_month = request.args.get('selected_month', today.strftime('%Y-%m'))

        try:
            year, month = map(int, selected_month.split('-'))
        except (ValueError, AttributeError):
            year, month = today.year, today.month

        start_date = date(year, month, 1)
        end_date = date(year, month, calendar.monthrange(year, month)[1])
        title_suffix = f"de {start_date.strftime('%B %Y')}"

    elif filter_type == 'year':
        selected_year = request.args.get('selected_year', str(today.year))
        try:
            year = int(selected_year)
            start_date = date(year, 1, 1)
            end_date = date(year, 12, 31)
            title_suffix = f"de l'année {year}"
        except:
            start_date = date(today.year, 1, 1)
            end_date = date(today.year, 12, 31)
            title_suffix = f"de l'année {today.year}"

    # Construire le filtre pour le type d'espace
    space_type_filters = []
    if space_type_filter != 'all' and space_type_filter:
        space_type_filters = [Booking.space_type == space_type_filter]

    # Créer des filtres de date basés sur les dates d'utilisation
    date_filters = [
        Booking.start_datetime <= datetime.combine(end_date, datetime.max.time()),
        Booking.end_datetime >= datetime.combine(start_date, datetime.min.time())
    ]
    filters = date_filters + space_type_filters

    # Calcul des statistiques pour la période actuelle
    # Calcul des statistiques pour la période actuelle avec daily_profit
    booking_stats_data = []
    for status in ['pending', 'confirmed', 'cancelled']:
        # Récupérer toutes les réservations pour ce statut
        status_bookings = Booking.query.filter(
            *filters,
            Booking.status == status
        ).all()
        
        # Calculer le daily_profit pour ces réservations
        status_daily_profit = 0
        for booking in status_bookings:
            # Déterminer la période d'intersection entre la réservation et la période filtrée
            booking_start = booking.start_datetime.date() if booking.start_datetime else start_date
            booking_end = booking.end_datetime.date() if booking.end_datetime else end_date
            
            # Trouver l'intersection entre la période de réservation et la période filtrée
            intersect_start = max(booking_start, start_date)
            intersect_end = min(booking_end, end_date)
            
            # Calculer le nombre de jours dans l'intersection
            if intersect_start <= intersect_end:
                days_in_period = (intersect_end - intersect_start).days + 1
                total_booking_days = (booking_end - booking_start).days + 1
                
                if booking.booking_type == 'monthly':
                    daily_rate = booking.total_price / total_booking_days
                    status_daily_profit += daily_rate * days_in_period
                elif booking.booking_type == 'daily':
                    daily_rate = booking.total_price / total_booking_days
                    status_daily_profit += daily_rate * days_in_period
                else:  # hourly
                    hourly_rate = booking.total_price / (total_booking_days * 24)
                    status_daily_profit += hourly_rate * (days_in_period * 24)
        
        booking_stats_data.append({
            'status': status,
            'count': len(status_bookings),
            'daily_profit': status_daily_profit
        })

    # Calculer les totaux
    total_bookings_period = sum(stat['count'] for stat in booking_stats_data) if booking_stats_data else 0
    total_daily_profit_period = sum(stat['daily_profit'] for stat in booking_stats_data) if booking_stats_data else 0.0

    # Préparation des données pour le template
    status_labels = {'pending': 'En attente', 'confirmed': 'Confirmée', 'cancelled': 'Annulée', 'completed': 'Terminée'}
    status_classes = {'pending': 'warning', 'confirmed': 'success', 'cancelled': 'danger', 'completed': 'info'}

    booking_stats = []
    for stat in booking_stats_data:
        percentage = round((stat['count'] / total_bookings_period) * 100, 1) if total_bookings_period > 0 else 0
        booking_stats.append({
            'status': stat['status'],
            'status_label': status_labels.get(stat['status'], stat['status']),
            'status_class': status_classes.get(stat['status'], 'secondary'),
            'count': stat['count'],
            'daily_profit': stat['daily_profit'],
            'percentage': percentage
        })

    # Calcul du taux de conversion et d'annulation (basé sur date_created)
    total_reservations_created = db.session.query(func.count(Booking.id)).filter(
        Booking.date_created >= datetime.combine(start_date, datetime.min.time()),
        Booking.date_created <= datetime.combine(end_date, datetime.max.time()),
        *space_type_filters
    ).scalar() or 0

    confirmed_reservations = db.session.query(func.count(Booking.id)).filter(
        Booking.date_created >= datetime.combine(start_date, datetime.min.time()),
        Booking.date_created <= datetime.combine(end_date, datetime.max.time()),
        Booking.status == 'confirmed',
        *space_type_filters
    ).scalar() or 0


    cancelled_reservations = db.session.query(func.count(Booking.id)).filter(
        Booking.date_created >= datetime.combine(start_date, datetime.min.time()),
        Booking.date_created <= datetime.combine(end_date, datetime.max.time()),
        Booking.status == 'cancelled',
        *space_type_filters
    ).scalar() or 0

    cancellation_rate = (cancelled_reservations / total_reservations_created * 100) if total_reservations_created > 0 else 0




    # Période précédente pour la comparaison
    if filter_type == 'day':
        prev_start_date = start_date - timedelta(days=1)
        prev_end_date = end_date - timedelta(days=1)
    elif filter_type == 'month':
        # Si le mois est janvier, la période précédente est décembre de l'année précédente
        if start_date.month == 1:
            prev_start_date = date(start_date.year - 1, 12, 1)
            prev_end_date = date(start_date.year - 1, 12, calendar.monthrange(start_date.year - 1, 12)[1])
        # Sinon, la période précédente est le mois précédent dans la même année
        else:
            prev_start_date = date(start_date.year, start_date.month - 1, 1)
            prev_end_date = date(start_date.year, start_date.month - 1, calendar.monthrange(start_date.year, start_date.month - 1)[1])

    elif filter_type == 'year':
        prev_start_date = start_date.replace(year=start_date.year - 1)
        prev_end_date = end_date.replace(year=end_date.year - 1)

    # Calcul des statistiques pour la période précédente
    prev_date_filters = [
        Booking.start_datetime <= datetime.combine(prev_end_date, datetime.max.time()),
        Booking.end_datetime >= datetime.combine(prev_start_date, datetime.min.time())
    ]
    prev_filters = prev_date_filters + space_type_filters

    prev_booking_stats_data = db.session.query(
        Booking.status,
        func.count(Booking.id).label('count'),
        func.sum(Booking.total_price).label('value')
    ).filter(*prev_filters).group_by(Booking.status).all()

    prev_total_bookings = sum(stat.count for stat in prev_booking_stats_data) if prev_booking_stats_data else 0
    # Calcul du profit journalier pour la période précédente
    prev_daily_profit = 0
    for booking in Booking.query.filter(*prev_filters, Booking.status == 'confirmed').all():
        # Déterminer la période d'intersection entre la réservation et la période précédente
        booking_start = booking.start_datetime.date() if booking.start_datetime else prev_start_date
        booking_end = booking.end_datetime.date() if booking.end_datetime else prev_end_date
        
        # Trouver l'intersection entre la période de réservation et la période précédente
        prev_intersect_start = max(booking_start, prev_start_date)
        prev_intersect_end = min(booking_end, prev_end_date)
        
        # Calculer le nombre de jours dans l'intersection
        if prev_intersect_start <= prev_intersect_end:
            prev_days_in_period = (prev_intersect_end - prev_intersect_start).days + 1
            prev_total_booking_days = (booking_end - booking_start).days + 1
            
            if booking.booking_type == 'monthly':
                # Pour les réservations mensuelles, calculer le prix journalier
                prev_daily_rate = booking.total_price / prev_total_booking_days
                prev_daily_profit += prev_daily_rate * prev_days_in_period
            elif booking.booking_type == 'daily':
                # Pour les réservations journalières, même calcul
                prev_daily_rate = booking.total_price / prev_total_booking_days
                prev_daily_profit += prev_daily_rate * prev_days_in_period
            else:  # hourly
                # Pour les réservations horaires, diviser par le nombre de jours de la période
                prev_hourly_rate = booking.total_price / (prev_total_booking_days * 24)
                prev_daily_profit += prev_hourly_rate * (prev_days_in_period * 24)
    # Calcul des variations


    # Calcul du profit journalier (répartition selon le type de réservation)
    # Calcul du profit journalier (répartition selon le type de réservation)
    daily_profit = 0
    for booking in Booking.query.filter(*filters, Booking.status == 'confirmed').all():
        # Déterminer la période d'intersection entre la réservation et la période filtrée
        booking_start = booking.start_datetime.date() if booking.start_datetime else start_date
        booking_end = booking.end_datetime.date() if booking.end_datetime else end_date
        
        # Trouver l'intersection entre la période de réservation et la période filtrée
        intersect_start = max(booking_start, start_date)
        intersect_end = min(booking_end, end_date)
        
        # Calculer le nombre de jours dans l'intersection
        if intersect_start <= intersect_end:
            days_in_period = (intersect_end - intersect_start).days + 1
            total_booking_days = (booking_end - booking_start).days + 1
            
            if booking.booking_type == 'monthly':
                # Pour les réservations mensuelles, calculer le prix journalier
                daily_rate = booking.total_price / total_booking_days
                daily_profit += daily_rate * days_in_period
            elif booking.booking_type == 'daily':
                # Pour les réservations journalières, même calcul
                daily_rate = booking.total_price / total_booking_days
                daily_profit += daily_rate * days_in_period
            else:  # hourly
                # Pour les réservations horaires, diviser par le nombre de jours de la période
                hourly_rate = booking.total_price / (total_booking_days * 24)  # Prix par heure
                daily_profit += hourly_rate * (days_in_period * 24)  # Convertir en équivalent journalier


    # Calcul des variations
    if prev_total_bookings > 0:
        bookings_variation = ((total_bookings_period - prev_total_bookings) / prev_total_bookings) * 100
    else:
        bookings_variation = 0

    # Calcul de la variation de revenu basée sur le profit journalier
    if prev_daily_profit > 0:
        revenue_variation = ((daily_profit - prev_daily_profit) / prev_daily_profit) * 100
    else:
        revenue_variation = 0

    # Statistiques par type d'espace
    # Statistiques par type d'espace (basées sur daily_profit)
    space_type_performance = []
    space_types = db.session.query(Booking.space_type.distinct()).filter(*filters, Booking.status == 'confirmed').all()

    for space_type in space_types:
        if space_type[0]:
            # Récupérer toutes les réservations pour ce type d'espace
            space_bookings = Booking.query.filter(
                *filters,
                Booking.status == 'confirmed',
                Booking.space_type == space_type[0]
            ).all()
            
            # Calculer le daily_profit pour ces réservations
            space_daily_profit = 0
            space_count = len(space_bookings)
            space_total_days = 0
            
            for booking in space_bookings:
                # Déterminer la période d'intersection entre la réservation et la période filtrée
                booking_start = booking.start_datetime.date() if booking.start_datetime else start_date
                booking_end = booking.end_datetime.date() if booking.end_datetime else end_date
                
                # Trouver l'intersection entre la période de réservation et la période filtrée
                intersect_start = max(booking_start, start_date)
                intersect_end = min(booking_end, end_date)
                
                # Calculer le nombre de jours dans l'intersection
                if intersect_start <= intersect_end:
                    days_in_period = (intersect_end - intersect_start).days + 1
                    total_booking_days = (booking_end - booking_start).days + 1
                    
                    if booking.booking_type == 'monthly':
                        daily_rate = booking.total_price / total_booking_days
                        space_daily_profit += daily_rate * days_in_period
                    elif booking.booking_type == 'daily':
                        daily_rate = booking.total_price / total_booking_days
                        space_daily_profit += daily_rate * days_in_period
                    else:  # hourly
                        hourly_rate = booking.total_price / (total_booking_days * 24)
                        space_daily_profit += hourly_rate * (days_in_period * 24)
                    
                    space_total_days += days_in_period
            
            # Calculer la durée moyenne
            avg_duration = space_total_days / space_count if space_count > 0 else 0
            
            space_type_performance.append({
                'space_type': space_type[0],
                'count': space_count,
                'daily_profit': space_daily_profit,
                'avg_duration': avg_duration
            })

    # Préparation des données pour les graphiques
    space_type_names = []
    revenue_data = []
    avg_duration_data = []
    space_type_performance_data = []
    for stat in space_type_performance:
        space_type = stat['space_type']
        if space_type == 'open':
            display_name = 'Open space'
        elif space_type == 'meeting':
            display_name = 'Salle de réunion'
        elif space_type == 'private':
            display_name = 'Bureau privé'
        else:
            display_name = space_type or 'Non spécifié'

        space_type_names.append(display_name)
        revenue_data.append(float(stat['daily_profit']) if stat['daily_profit'] else 0)
        avg_duration_data.append(float(stat['avg_duration']) if stat['avg_duration'] else 0)
        space_type_performance_data.append({
            'space_type': display_name,
            'count': stat['count'],
            'daily_profit': stat['daily_profit'] or 0,
            'avg_duration': stat['avg_duration'] or 0
        })



    # Graphique principal (réservations par heure/jour/mois)
    if filter_type == 'day':
        dates = [f"{hour:02d}:00" for hour in range(24)]
        booking_counts = [0] * 24

        # Utiliser les mêmes filtres que pour les réservations récentes
        day_bookings = Booking.query.filter(
            *filters,  # Utilise les filtres déjà définis (inclut la période et le type d'espace)
            Booking.status == 'confirmed'
        ).all()

        print(f"Nombre de réservations trouvées pour le jour : {len(day_bookings)}")

        for booking in day_bookings:
            # Ignorer les réservations sans dates
            if not booking.start_datetime or not booking.end_datetime:
                continue

            # Limiter la réservation à la journée sélectionnée
            booking_start = max(booking.start_datetime, datetime.combine(start_date, datetime.min.time()))
            booking_end = min(booking.end_datetime, datetime.combine(start_date, datetime.max.time()))

            # Pour chaque heure de la journée
            for hour in range(24):
                hour_start = datetime.combine(start_date, time(hour, 0))
                hour_end = datetime.combine(start_date, time(hour, 59, 59))

                # Si la réservation chevauche cette heure
                if booking_start <= hour_end and booking_end >= hour_start:
                    booking_counts[hour] += 1

        print("Booking counts après calcul :", booking_counts)






 




    elif filter_type == 'month':
        dates = []
        booking_counts = []
        current = start_date
        while current <= end_date:
            dates.append(current.strftime('%d/%m'))
            day_start = datetime.combine(current, datetime.min.time())
            day_end = datetime.combine(current, datetime.max.time())
            count = db.session.query(func.count(Booking.id)).filter(
                Booking.start_datetime <= day_end,
                Booking.end_datetime >= day_start,
                Booking.status == 'confirmed',
                *space_type_filters
            ).scalar() or 0
            booking_counts.append(count)
            current += timedelta(days=1)

    else:  # year
        dates = []
        booking_counts = []
        for month in range(1, 13):
            month_start = date(start_date.year, month, 1)
            month_end = date(start_date.year, month, calendar.monthrange(start_date.year, month)[1])
            if month_start <= end_date:
                dates.append(month_start.strftime('%b'))
                count = db.session.query(func.count(Booking.id)).filter(
                    or_(
                        and_(
                            Booking.start_datetime >= datetime.combine(month_start, datetime.min.time()),
                            Booking.start_datetime <= datetime.combine(month_end, datetime.max.time())
                        ),
                        and_(
                            Booking.end_datetime >= datetime.combine(month_start, datetime.min.time()),
                            Booking.end_datetime <= datetime.combine(month_end, datetime.max.time())
                        ),
                        and_(
                            Booking.start_datetime <= datetime.combine(month_start, datetime.min.time()),
                            Booking.end_datetime >= datetime.combine(month_end, datetime.max.time())
                        )
                    ),
                    Booking.status == 'confirmed',
                    *space_type_filters
                ).scalar() or 0
                booking_counts.append(count)

    # Réservations récentes
    recent_bookings = Booking.query.filter(
        *filters  # Utilise les mêmes filtres que pour le reste des statistiques
    ).order_by(Booking.date_created.desc()).all()

    # Messages (indépendants de la période)
    unread_messages = db.session.query(func.count(ContactMessage.id)).filter(
        ContactMessage.status == 'unread'
    ).scalar() or 0

    message_stats_data = db.session.query(
        ContactMessage.status,
        func.count(ContactMessage.id).label('count')
    ).group_by(ContactMessage.status).all()

    message_stats = []
    message_status_labels = {'unread': 'Non lu', 'read': 'Lu', 'replied': 'Répondu'}
    message_status_classes = {'unread': 'warning', 'read': 'success', 'replied': 'info'}

    for stat in message_stats_data:
        message_stats.append({
            'status': stat.status,
            'status_label': message_status_labels.get(stat.status, stat.status),
            'status_class': message_status_classes.get(stat.status, 'secondary'),
            'count': stat.count
        })

    # Préparation des données pour le template
    status_labels = {'pending': 'En attente', 'confirmed': 'Confirmée', 'cancelled': 'Annulée'}
    status_classes = {'pending': 'warning', 'confirmed': 'success', 'cancelled': 'danger'}

    booking_stats = []
    for stat in booking_stats_data:
        percentage = round((stat['count'] / total_bookings_period) * 100, 1) if total_bookings_period > 0 else 0
        booking_stats.append({
            'status': stat['status'],
            'status_label': status_labels.get(stat['status'], stat['status']),
            'status_class': status_classes.get(stat['status'], 'secondary'),
            'count': stat['count'],
            'daily_profit': stat['daily_profit'],
            'percentage': percentage
        })

    # Formatage des réservations récentes
    formatted_recent_bookings = []
    for booking in recent_bookings:
        duration_str = "N/A"
        if booking.start_datetime and booking.end_datetime:
            delta = booking.end_datetime - booking.start_datetime
            if booking.booking_type == 'monthly':
                num_months = (booking.end_datetime.year - booking.start_datetime.year) * 12 + (booking.end_datetime.month - booking.start_datetime.month)
                duration_str = f"{num_months} mois"
            elif booking.booking_type == 'daily':
                num_days = delta.days + 1
                duration_str = f"{num_days} jours"
            elif booking.booking_type == 'hourly':
                num_hours = delta.total_seconds() / 3600
                duration_str = f"{int(num_hours)} heures"

        formatted_recent_bookings.append({
            'id': booking.id,
            'reference': f"RES{booking.id:06d}",
            'client_name': booking.full_name,
            'space_type': booking.space_type,
            'booking_type': booking.booking_type,
            'booking_date': booking.start_datetime.strftime('%d/%m/%Y') if booking.start_datetime else 'N/A',
            'amount': booking.total_price,
            'status': status_labels.get(booking.status, booking.status),
            'status_class': status_classes.get(booking.status, 'secondary'),
            'duration': duration_str,  # <-- Ajoutez cette ligne
        })


    # Préparation des options pour les sélecteurs
    current_year = datetime.now().year
    years = [str(year) for year in range(2023, 2041)]
    months = []
    for month in range(1, 13):
        month_date = date(current_year, month, 1)
        months.append({
            'value': f"{current_year}-{month:02d}",
            'label': month_date.strftime('%B %Y')
        })


    # Statistiques par type de réservation (pour le graphique)
    # Statistiques par type de réservation (pour le graphique) avec daily_profit
    booking_type_stats_data = []
    booking_types = db.session.query(Booking.booking_type.distinct()).filter(*filters, Booking.status == 'confirmed').all()

    for booking_type in booking_types:
        if booking_type[0]:
            # Récupérer toutes les réservations pour ce type
            type_bookings = Booking.query.filter(
                *filters,
                Booking.status == 'confirmed',
                Booking.booking_type == booking_type[0]
            ).all()
            
            # Calculer le daily_profit pour ces réservations
            type_daily_profit = 0
            
            for booking in type_bookings:
                # Déterminer la période d'intersection entre la réservation et la période filtrée
                booking_start = booking.start_datetime.date() if booking.start_datetime else start_date
                booking_end = booking.end_datetime.date() if booking.end_datetime else end_date
                
                # Trouver l'intersection entre la période de réservation et la période filtrée
                intersect_start = max(booking_start, start_date)
                intersect_end = min(booking_end, end_date)
                
                # Calculer le nombre de jours dans l'intersection
                if intersect_start <= intersect_end:
                    days_in_period = (intersect_end - intersect_start).days + 1
                    total_booking_days = (booking_end - booking_start).days + 1
                    
                    if booking.booking_type == 'monthly':
                        daily_rate = booking.total_price / total_booking_days
                        type_daily_profit += daily_rate * days_in_period
                    elif booking.booking_type == 'daily':
                        daily_rate = booking.total_price / total_booking_days
                        type_daily_profit += daily_rate * days_in_period
                    else:  # hourly
                        hourly_rate = booking.total_price / (total_booking_days * 24)
                        type_daily_profit += hourly_rate * (days_in_period * 24)
            
            booking_type_stats_data.append({
                'booking_type': booking_type[0],
                'count': len(type_bookings),
                'daily_profit': type_daily_profit
            })

    # Préparation pour le graphique
    booking_type_names = []
    booking_type_counts = [] 
    month_names = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
] # Changé de 'counts' à 'profits'


    for stat in booking_type_stats_data:
        booking_type = stat['booking_type']
        if booking_type == 'monthly':
            display_name = 'Par Mois'
        elif booking_type == 'daily':
            display_name = 'Par Jour'
        elif booking_type == 'hourly':
            display_name = 'Par Heure'
        else:
            display_name = booking_type or 'Non spécifié'

        booking_type_names.append(display_name)
        booking_type_counts.append(stat['count'])


    return render_template(
        'admin_dashboard.html',
        active_tab='admin_stats',
        filter_type=filter_type,
        selected_date=selected_date if filter_type == 'day' else today.strftime('%Y-%m-%d'),
        selected_month=selected_month if filter_type == 'month' else today.strftime('%Y-%m'),
        selected_year=selected_year if filter_type == 'year' else str(current_year),
        title_suffix=title_suffix,
        booking_stats=booking_stats,
        total_bookings_period=total_bookings_period,
        total_value_period=total_daily_profit_period,
        daily_profit=round(daily_profit, 2),
        cancellation_rate=round(cancellation_rate, 1),
        comparison_data={
            'period': {
                'bookings': total_bookings_period,
                'revenue': total_daily_profit_period,
                'daily_profit': daily_profit  # Ajout pour clarté

            },
            'previous_period': {
                'bookings': prev_total_bookings,
                'revenue': prev_daily_profit,
                'daily_profit': prev_daily_profit  # Ajout pour clarté

            },
            'variations': {
                'bookings': bookings_variation,
                'revenue': revenue_variation
            }
        },
        prev_period_label=f"Période précédente ({prev_start_date.strftime('%d/%m/%Y') if filter_type == 'day' else prev_start_date.strftime('%B %Y') if filter_type == 'month' else prev_start_date.year})",
        space_type_names=space_type_names,
        revenue_data=revenue_data,
        avg_duration_data=avg_duration_data,
        space_type_performance_data=space_type_performance_data,
        dates=dates,
        booking_counts=booking_counts,
        recent_bookings=formatted_recent_bookings,
        years=years,
        months=months,
        current_time=datetime.now(),
        available_space_types=available_space_types,
        selected_space_type=space_type_filter,
        unread_messages=unread_messages,
        message_stats=message_stats,
        booking_type_names=booking_type_names,
        booking_type_counts=booking_type_counts,
        datetime=datetime, 
        str=str


    )










def generate_styled_excel(data, filename="export.xlsx"):
    df = pd.DataFrame(data)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Données')
        workbook = writer.book
        worksheet = writer.sheets['Données']

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alignment = Alignment(horizontal='center', vertical='center')

        # Appliquer les styles aux en-têtes
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment

        # Appliquer les bordures et l'alignement à toutes les cellules
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

        # Ajuster la largeur des colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
@admin_stats.route("/dashboard/admin_stats/export_excel")
@login_required
def export_excel():
    # Récupérer les paramètres de filtre
    filter_type = request.args.get('filter_type')
    selected_date = request.args.get('selected_date')
    selected_month = request.args.get('selected_month')
    selected_year = request.args.get('selected_year')
    space_type = request.args.get('space_type')

    # Initialiser la requête
    query = Booking.query

    # Définir les dates de début et de fin de la période filtrée
    start_period = None
    end_period = None

    if filter_type == 'day' and selected_date:
        selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        start_period = datetime.combine(selected_date, datetime.min.time())
        end_period = datetime.combine(selected_date, datetime.max.time())
        query = query.filter(
            db.or_(
                db.and_(
                    Booking.start_datetime <= end_period,
                    Booking.end_datetime >= start_period
                ),
                db.and_(
                    Booking.start_datetime >= start_period,
                    Booking.start_datetime <= end_period
                )
            )
        )
    elif filter_type == 'month' and selected_month:
        year, month = map(int, selected_month.split('-'))
        if month == 12:
            end_year = year + 1
            end_month = 1
        else:
            end_year = year
            end_month = month + 1
        start_period = datetime(year, month, 1)
        end_period = datetime(end_year, end_month, 1) - timedelta(days=1)
        query = query.filter(
            db.or_(
                db.and_(
                    Booking.start_datetime <= end_period,
                    Booking.end_datetime >= start_period
                ),
                db.and_(
                    Booking.start_datetime >= start_period,
                    Booking.start_datetime <= end_period
                )
            )
        )
    elif filter_type == 'year' and selected_year:
        start_period = datetime(int(selected_year), 1, 1)
        end_period = datetime(int(selected_year), 12, 31, 23, 59, 59)
        query = query.filter(
            db.or_(
                db.and_(
                    Booking.start_datetime <= end_period,
                    Booking.end_datetime >= start_period
                ),
                db.and_(
                    Booking.start_datetime >= start_period,
                    Booking.start_datetime <= end_period
                )
            )
        )

    if space_type and space_type != 'all':
        query = query.filter(Booking.space_type == space_type)

    # Récupérer les réservations
    bookings = query.order_by(Booking.start_datetime).all()

    # Préparer les données pour l'export
    data = []
    for booking in bookings:
        booking_type = booking.booking_type
        # Changer l'affichage des types de réservation
        if booking_type == 'monthly':
            display_booking_type = 'Par Mois'
        elif booking_type == 'daily':
            display_booking_type = 'Par Jour'
        elif booking_type == 'hourly':
            display_booking_type = 'Par Heure'
        else:
            display_booking_type = booking_type or 'Non spécifié'
        space_type = booking.space_type
        # Changer l'affichage des types d'espace
        if space_type == 'open':
            display_space_type = 'open space'
        elif space_type == 'meeting':
            display_space_type = 'salle de réunion'
        elif space_type == 'private':
            display_space_type = 'bureau privé'
        else:
            display_space_type = space_type or 'Non spécifié'
        data.append({
            "ID": booking.id,
            "Référence": f"RES{booking.id:06d}",
            "Numéro d'espace": booking.space_number or "N/A",
            "Type de réservation": display_booking_type,
            "Type d'espace": display_space_type,
            "Date de début": booking.start_datetime.strftime('%d/%m/%Y %H:%M') if booking.start_datetime else '',
            "Date de fin": booking.end_datetime.strftime('%d/%m/%Y %H:%M') if booking.end_datetime else '',
            "Durée": f"{booking.duration} {'heure(s)' if booking.booking_type == 'hourly' else 'jour(s)' if booking.booking_type == 'daily' else 'mois'}",
            "Client": booking.full_name,
            "Email": booking.email,
            "Téléphone": booking.phone,
            "Entreprise": booking.company or "N/A",
            "Demandes spéciales": booking.special_requests or "Aucune",
            "Méthode de paiement": booking.payment_method,
            "Montant total": f"{booking.total_price} DH",
            "Statut": booking.status,
            "Créé par": booking.created_by,
        })

    return generate_styled_excel(data, filename="statistiques_export.xlsx")
